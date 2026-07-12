import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Crear un nuevo libro de trabajo
wb = openpyxl.Workbook()

# Estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
result_font = Font(bold=True)
result_fill = PatternFill("solid", fgColor="E2EFDA")
title_font = Font(bold=True, size=14)

# ---------------------------------------------------------
# Hoja 1: Capacidad en Arcilla
# ---------------------------------------------------------
ws_arcilla = wb.active
ws_arcilla.title = "Arcilla (Método Alfa)"

# Título
ws_arcilla.merge_cells('A1:C1')
ws_arcilla['A1'] = "Cálculo de Capacidad en Arcilla"
ws_arcilla['A1'].font = title_font

# Entradas
ws_arcilla['A3'] = "Datos de Entrada"
ws_arcilla['A3'].font = header_font
ws_arcilla['A3'].fill = header_fill
ws_arcilla['B3'] = "Valor"
ws_arcilla['B3'].font = header_font
ws_arcilla['B3'].fill = header_fill
ws_arcilla['C3'] = "Unidad"
ws_arcilla['C3'].font = header_font
ws_arcilla['C3'].fill = header_fill

entradas_arcilla = [
    ("Diámetro del pilote (D)", 0.4, "m"),
    ("Longitud del pilote (L)", 15.0, "m"),
    ("Cohesión no drenada (cu)", 50.0, "kN/m2"),
    ("Factor de adhesión (alfa)", 0.7, "-")
]

for i, (param, valor, unidad) in enumerate(entradas_arcilla, start=4):
    ws_arcilla.cell(row=i, column=1, value=param)
    ws_arcilla.cell(row=i, column=2, value=valor)
    ws_arcilla.cell(row=i, column=3, value=unidad)

# Cálculos Intermedios
ws_arcilla['A9'] = "Cálculos Intermedios"
ws_arcilla['A9'].font = header_font
ws_arcilla['A9'].fill = header_fill

ws_arcilla['A10'] = "Área de la punta (Ap)"
ws_arcilla['B10'] = "=PI()*(B4^2)/4"
ws_arcilla['C10'] = "m2"

ws_arcilla['A11'] = "Perímetro"
ws_arcilla['B11'] = "=PI()*B4"
ws_arcilla['C11'] = "m"

ws_arcilla['A12'] = "Factor Nc*"
ws_arcilla['B12'] = 9.0
ws_arcilla['C12'] = "-"

# Resultados
ws_arcilla['A14'] = "Resultados"
ws_arcilla['A14'].font = header_font
ws_arcilla['A14'].fill = header_fill

ws_arcilla['A15'] = "Resistencia de punta (Qp)"
ws_arcilla['B15'] = "=B10 * B6 * B12"
ws_arcilla['C15'] = "kN"
ws_arcilla['A15'].fill = result_fill
ws_arcilla['B15'].fill = result_fill
ws_arcilla['C15'].fill = result_fill

ws_arcilla['A16'] = "Fricción lateral (Qs)"
ws_arcilla['B16'] = "=B11 * B5 * B7 * B6"
ws_arcilla['C16'] = "kN"
ws_arcilla['A16'].fill = result_fill
ws_arcilla['B16'].fill = result_fill
ws_arcilla['C16'].fill = result_fill

ws_arcilla['A17'] = "Capacidad última total (Qu)"
ws_arcilla['B17'] = "=B15 + B16"
ws_arcilla['C17'] = "kN"
ws_arcilla['A17'].font = result_font
ws_arcilla['B17'].font = result_font
ws_arcilla['C17'].font = result_font
ws_arcilla['A17'].fill = result_fill
ws_arcilla['B17'].fill = result_fill
ws_arcilla['C17'].fill = result_fill

# Ajustar ancho columnas
ws_arcilla.column_dimensions['A'].width = 30
ws_arcilla.column_dimensions['B'].width = 15
ws_arcilla.column_dimensions['C'].width = 10


# ---------------------------------------------------------
# Hoja 2: Fórmula ENR Modificada
# ---------------------------------------------------------
ws_enr = wb.create_sheet(title="Fórmula ENR Modificada")

# Título
ws_enr.merge_cells('A1:C1')
ws_enr['A1'] = "Cálculo con Fórmula Dinámica de Hincado ENR"
ws_enr['A1'].font = title_font

# Entradas
ws_enr['A3'] = "Datos del Martinete e Hincado"
ws_enr['A3'].font = header_font
ws_enr['A3'].fill = header_fill
ws_enr['B3'] = "Valor"
ws_enr['B3'].font = header_font
ws_enr['B3'].fill = header_fill
ws_enr['C3'] = "Unidad"
ws_enr['C3'].font = header_font
ws_enr['C3'].fill = header_fill

entradas_enr = [
    ("Eficiencia del martinete (E)", 0.8, "-"),
    ("Peso del ariete (W_R)", 33.36, "kN"),
    ("Altura de caída (h)", 1250, "mm"),
    ("Penetración por golpe (S)", 3.175, "mm"),
    ("Constante (C)", 2.54, "mm"),
    ("Coeficiente de restitución (n)", 0.4, "-"),
    ("Peso del pilote (W_p)", 55.95, "kN")
]

for i, (param, valor, unidad) in enumerate(entradas_enr, start=4):
    ws_enr.cell(row=i, column=1, value=param)
    ws_enr.cell(row=i, column=2, value=valor)
    ws_enr.cell(row=i, column=3, value=unidad)

# Cálculos Intermedios
ws_enr['A12'] = "Cálculos Intermedios"
ws_enr['A12'].font = header_font
ws_enr['A12'].fill = header_fill

ws_enr['A13'] = "Energía Neta Entregada"
ws_enr['B13'] = "=B4 * B5 * B6"
ws_enr['C13'] = "kN-mm"

ws_enr['A14'] = "Término de Pesos"
ws_enr['B14'] = "=(B5 + (B9^2)*B10) / (B5 + B10)"
ws_enr['C14'] = "-"

# Resultados
ws_enr['A16'] = "Resultados"
ws_enr['A16'].font = header_font
ws_enr['A16'].fill = header_fill

ws_enr['A17'] = "Capacidad última total (Qu)"
ws_enr['B17'] = "=(B13 / (B7 + B8)) * B14"
ws_enr['C17'] = "kN"
ws_enr['A17'].font = result_font
ws_enr['B17'].font = result_font
ws_enr['C17'].font = result_font
ws_enr['A17'].fill = result_fill
ws_enr['B17'].fill = result_fill
ws_enr['C17'].fill = result_fill

# Ajustar ancho columnas
ws_enr.column_dimensions['A'].width = 40
ws_enr.column_dimensions['B'].width = 15
ws_enr.column_dimensions['C'].width = 10


# ---------------------------------------------------------
# Hoja 3: Arena (SPT Meyerhof)
# ---------------------------------------------------------
ws_spt = wb.create_sheet(title="Arena (SPT Meyerhof)")

# Título
ws_spt.merge_cells('A1:C1')
ws_spt['A1'] = "Cálculo de Capacidad en Arena usando SPT"
ws_spt['A1'].font = title_font

# Entradas
ws_spt['A3'] = "Datos de Entrada"
ws_spt['A3'].font = header_font
ws_spt['A3'].fill = header_fill
ws_spt['B3'] = "Valor"
ws_spt['B3'].font = header_font
ws_spt['B3'].fill = header_fill
ws_spt['C3'] = "Unidad"
ws_spt['C3'].font = header_font
ws_spt['C3'].fill = header_fill

entradas_spt = [
    ("Diámetro del pilote (D)", 0.4, "m"),
    ("Longitud del pilote (L)", 15.0, "m"),
    ("Presión atmosférica (pa)", 100.0, "kN/m2"),
    ("N60 promedio", 25.0, "-")
]

for i, (param, valor, unidad) in enumerate(entradas_spt, start=4):
    ws_spt.cell(row=i, column=1, value=param)
    ws_spt.cell(row=i, column=2, value=valor)
    ws_spt.cell(row=i, column=3, value=unidad)

# Cálculos Intermedios
ws_spt['A9'] = "Cálculos Intermedios"
ws_spt['A9'].font = header_font
ws_spt['A9'].fill = header_fill

ws_spt['A10'] = "Área de la punta (Ap)"
ws_spt['B10'] = "=PI()*(B4^2)/4"
ws_spt['C10'] = "m2"

ws_spt['A11'] = "Perímetro"
ws_spt['B11'] = "=PI()*B4"
ws_spt['C11'] = "m"

ws_spt['A12'] = "L/D"
ws_spt['B12'] = "=B5/B4"
ws_spt['C12'] = "-"

ws_spt['A13'] = "qp calculado"
ws_spt['B13'] = "=0.4*B6*B7*B12"
ws_spt['C13'] = "kN/m2"

ws_spt['A14'] = "qp límite"
ws_spt['B14'] = "=4*B6*B7"
ws_spt['C14'] = "kN/m2"

ws_spt['A15'] = "qp definitivo"
ws_spt['B15'] = "=MIN(B13, B14)"
ws_spt['C15'] = "kN/m2"

ws_spt['A16'] = "Fricción unitaria (f)"
ws_spt['B16'] = "=0.02*B6*B7"
ws_spt['C16'] = "kN/m2"

# Resultados
ws_spt['A18'] = "Resultados"
ws_spt['A18'].font = header_font
ws_spt['A18'].fill = header_fill

ws_spt['A19'] = "Resistencia de punta (Qp)"
ws_spt['B19'] = "=B10 * B15"
ws_spt['C19'] = "kN"
ws_spt['A19'].fill = result_fill
ws_spt['B19'].fill = result_fill
ws_spt['C19'].fill = result_fill

ws_spt['A20'] = "Fricción lateral (Qs)"
ws_spt['B20'] = "=B11 * B5 * B16"
ws_spt['C20'] = "kN"
ws_spt['A20'].fill = result_fill
ws_spt['B20'].fill = result_fill
ws_spt['C20'].fill = result_fill

ws_spt['A21'] = "Capacidad última total (Qu)"
ws_spt['B21'] = "=B19 + B20"
ws_spt['C21'] = "kN"
ws_spt['A21'].font = result_font
ws_spt['B21'].font = result_font
ws_spt['C21'].font = result_font
ws_spt['A21'].fill = result_fill
ws_spt['B21'].fill = result_fill
ws_spt['C21'].fill = result_fill

# Ajustar ancho columnas
ws_spt.column_dimensions['A'].width = 30
ws_spt.column_dimensions['B'].width = 15
ws_spt.column_dimensions['C'].width = 10

# Guardar
wb.save("Calculo_Pilotes_v2.xlsx")
print("Archivo Calculo_Pilotes_v2.xlsx actualizado correctamente con la nueva hoja.")
